import csv
from datetime import date, datetime
from io import BytesIO, StringIO
import json
import re
import sqlite3
from typing import Any
from zipfile import BadZipFile

from fastapi import status
from openpyxl import load_workbook
from pydantic import ValidationError

from ..core.exceptions import AppError
from ..schemas.review import (
    ReviewImportItem,
    ReviewImportRequest,
    ReviewImportResult,
    ReviewUploadFormat,
    ReviewUploadImportRequest,
    ReviewUploadPreprocessingSummary,
)
from .review_ingestion import ReviewIngestionService


class ReviewUploadImportService:
    CANONICAL_COLUMNS = (
        "source_review_id",
        "reviewer_name",
        "rating",
        "language",
        "review_text",
        "review_created_at",
        "status",
        "response_status",
        "source_metadata",
    )
    REQUIRED_COLUMNS = ("source_review_id", "review_text")
    DELIMITERS = (",", "\t", "|", ";")
    FORMAT_BY_EXTENSION = {
        ".csv": ReviewUploadFormat.CSV,
        ".xlsx": ReviewUploadFormat.XLSX,
        ".json": ReviewUploadFormat.JSON,
        ".txt": ReviewUploadFormat.TXT,
        ".db": ReviewUploadFormat.SQLITE,
        ".sqlite": ReviewUploadFormat.SQLITE,
    }
    COLUMN_ALIASES = {
        "source_review_id": [
            "review_id",
            "external_id",
            "external_review_id",
            "source_id",
            "record_id",
        ],
        "reviewer_name": [
            "reviewer",
            "author",
            "customer",
            "customer_name",
            "user_name",
            "username",
            "name",
        ],
        "rating": [
            "star_rating",
            "stars",
            "score",
            "review_rating",
            "recommendation_rating",
        ],
        "language": [
            "language_code",
            "lang",
            "locale",
        ],
        "review_text": [
            "comment",
            "review",
            "text",
            "body",
            "message",
            "content",
            "feedback",
        ],
        "review_created_at": [
            "created_at",
            "created_time",
            "created_date",
            "review_date",
            "submitted_at",
            "timestamp",
            "date",
        ],
        "status": [
            "review_status",
        ],
        "response_status": [
            "reply_status",
            "response_state",
        ],
        "source_metadata": [
            "metadata",
            "meta",
            "original_payload",
        ],
    }

    def __init__(
        self,
        *,
        review_ingestion_service: ReviewIngestionService,
    ) -> None:
        self.review_ingestion_service = review_ingestion_service

    def import_reviews(
        self,
        *,
        upload: ReviewUploadImportRequest,
        filename: str,
        content_type: str | None,
        content: bytes,
    ) -> ReviewImportResult:
        detected_format = self._detect_format(filename, content_type)
        extracted_rows, selected_source_name = self._extract_rows(
            detected_format=detected_format,
            filename=filename,
            content=content,
        )
        header_mapping = self._build_header_mapping(extracted_rows)
        staged_rows = self._build_canonical_rows(extracted_rows, header_mapping)
        csv_payload = self._stage_rows_as_csv(staged_rows)
        reviews = self._parse_canonical_csv(csv_payload)
        import_payload = ReviewImportRequest(
            business_id=upload.business_id,
            review_source_id=upload.review_source_id,
            source=upload.source,
            reviews=reviews,
        )
        summary = ReviewUploadPreprocessingSummary(
            original_filename=filename,
            detected_format=detected_format,
            source_row_count=len(staged_rows),
            mapped_columns=header_mapping,
            selected_source_name=selected_source_name,
        )
        result, _ = self.review_ingestion_service.import_reviews(
            import_payload,
            input_reference_extra={
                "upload": summary.model_dump(mode="json"),
            },
        )
        return result

    def _detect_format(
        self, filename: str, content_type: str | None
    ) -> ReviewUploadFormat:
        lowered_name = filename.lower()
        for extension, detected_format in self.FORMAT_BY_EXTENSION.items():
            if lowered_name.endswith(extension):
                return detected_format

        normalized_content_type = (content_type or "").lower()
        if "csv" in normalized_content_type:
            return ReviewUploadFormat.CSV
        if "spreadsheetml" in normalized_content_type or "excel" in normalized_content_type:
            return ReviewUploadFormat.XLSX
        if "json" in normalized_content_type:
            return ReviewUploadFormat.JSON
        if normalized_content_type.startswith("text/"):
            return ReviewUploadFormat.TXT

        raise AppError(
            code="UNSUPPORTED_UPLOAD_FORMAT",
            message="Uploaded file format is not supported.",
            status_code=status.HTTP_400_BAD_REQUEST,
            details={"filename": filename, "content_type": content_type},
        )

    def _extract_rows(
        self,
        *,
        detected_format: ReviewUploadFormat,
        filename: str,
        content: bytes,
    ) -> tuple[list[dict[str, Any]], str | None]:
        if not content:
            raise self._parse_error(
                message="Uploaded file is empty.",
                filename=filename,
                detected_format=detected_format,
            )

        if detected_format in {ReviewUploadFormat.CSV, ReviewUploadFormat.TXT}:
            return self._extract_delimited_rows(
                content=content,
                filename=filename,
                detected_format=detected_format,
            )
        if detected_format == ReviewUploadFormat.JSON:
            return self._extract_json_rows(content=content, filename=filename)
        if detected_format == ReviewUploadFormat.XLSX:
            return self._extract_xlsx_rows(content=content, filename=filename)
        if detected_format == ReviewUploadFormat.SQLITE:
            return self._extract_sqlite_rows(content=content, filename=filename)

        raise self._parse_error(
            message="Uploaded file format is not supported.",
            filename=filename,
            detected_format=detected_format,
        )

    def _extract_delimited_rows(
        self,
        *,
        content: bytes,
        filename: str,
        detected_format: ReviewUploadFormat,
    ) -> tuple[list[dict[str, Any]], str | None]:
        text = self._decode_text(content)
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters="".join(self.DELIMITERS))
        except csv.Error as exc:
            raise self._parse_error(
                message="Could not detect a stable delimiter for the uploaded text.",
                filename=filename,
                detected_format=detected_format,
                details={"supported_delimiters": list(self.DELIMITERS)},
            ) from exc

        reader = csv.DictReader(StringIO(text), dialect=dialect)
        if not reader.fieldnames:
            raise self._parse_error(
                message="Uploaded delimited text must include a header row.",
                filename=filename,
                detected_format=detected_format,
            )

        rows = [
            row
            for row in reader
            if row and any(self._has_value(value) for value in row.values())
        ]
        if not rows:
            raise self._parse_error(
                message="Uploaded file did not contain any data rows.",
                filename=filename,
                detected_format=detected_format,
            )
        return rows, None

    def _extract_json_rows(
        self, *, content: bytes, filename: str
    ) -> tuple[list[dict[str, Any]], str | None]:
        try:
            parsed = json.loads(self._decode_text(content))
        except json.JSONDecodeError as exc:
            raise self._parse_error(
                message="Uploaded JSON could not be parsed.",
                filename=filename,
                detected_format=ReviewUploadFormat.JSON,
            ) from exc

        if isinstance(parsed, list):
            raw_rows = parsed
        elif isinstance(parsed, dict) and isinstance(parsed.get("reviews"), list):
            raw_rows = parsed["reviews"]
        else:
            raise self._parse_error(
                message="Uploaded JSON must be an array or contain a top-level 'reviews' array.",
                filename=filename,
                detected_format=ReviewUploadFormat.JSON,
            )

        rows = [row for row in raw_rows if isinstance(row, dict)]
        if len(rows) != len(raw_rows):
            raise self._parse_error(
                message="Uploaded JSON rows must be objects.",
                filename=filename,
                detected_format=ReviewUploadFormat.JSON,
            )
        if not rows:
            raise self._parse_error(
                message="Uploaded JSON did not contain any review rows.",
                filename=filename,
                detected_format=ReviewUploadFormat.JSON,
            )
        return rows, None

    def _extract_xlsx_rows(
        self, *, content: bytes, filename: str
    ) -> tuple[list[dict[str, Any]], str | None]:
        try:
            workbook = load_workbook(
                filename=BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except (BadZipFile, OSError, ValueError) as exc:
            raise self._parse_error(
                message="Uploaded XLSX file could not be parsed.",
                filename=filename,
                detected_format=ReviewUploadFormat.XLSX,
            ) from exc

        non_empty_sheets: list[str] = []
        selected_rows: list[dict[str, Any]] = []
        selected_sheet_name: str | None = None
        try:
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = list(worksheet.iter_rows(values_only=True))
                if not any(
                    any(self._has_value(cell) for cell in row)
                    for row in rows
                ):
                    continue
                non_empty_sheets.append(sheet_name)
                if len(non_empty_sheets) > 1:
                    raise AppError(
                        code="UPLOAD_SOURCE_SELECTION_ERROR",
                        message="Uploaded workbook must contain exactly one non-empty worksheet.",
                        status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                        details={"candidate_sources": non_empty_sheets},
                    )
                selected_sheet_name = sheet_name
                selected_rows = self._rows_from_tabular_values(rows)
        finally:
            workbook.close()

        if not non_empty_sheets:
            raise self._parse_error(
                message="Uploaded workbook did not contain any non-empty worksheets.",
                filename=filename,
                detected_format=ReviewUploadFormat.XLSX,
            )
        if not selected_rows:
            raise self._parse_error(
                message="Selected worksheet did not contain any data rows.",
                filename=filename,
                detected_format=ReviewUploadFormat.XLSX,
            )
        return selected_rows, selected_sheet_name

    def _extract_sqlite_rows(
        self, *, content: bytes, filename: str
    ) -> tuple[list[dict[str, Any]], str | None]:
        connection = sqlite3.connect(":memory:")
        try:
            connection.deserialize(content)
            cursor = connection.cursor()
            table_names = [
                row[0]
                for row in cursor.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
                ).fetchall()
            ]
            candidates: list[tuple[str, dict[str, str]]] = []
            for table_name in table_names:
                columns = [
                    row[1]
                    for row in cursor.execute(f'PRAGMA table_info("{table_name}")').fetchall()
                ]
                mapping = self._build_header_mapping([dict.fromkeys(columns, None)])
                if all(required in mapping for required in self.REQUIRED_COLUMNS):
                    candidates.append((table_name, mapping))

            if len(candidates) > 1:
                raise AppError(
                    code="UPLOAD_SOURCE_SELECTION_ERROR",
                    message="Uploaded SQLite database must contain exactly one review-like table.",
                    status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                    details={"candidate_sources": [table_name for table_name, _ in candidates]},
                )
            if not candidates:
                raise AppError(
                    code="UPLOAD_COLUMN_MAPPING_ERROR",
                    message="Uploaded SQLite database did not contain a recognizable review table.",
                    status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                    details={"detected_tables": table_names, "missing_columns": list(self.REQUIRED_COLUMNS)},
                )

            table_name, _ = candidates[0]
            cursor.execute(f'SELECT * FROM "{table_name}"')
            column_names = [description[0] for description in cursor.description or []]
            rows = [
                dict(zip(column_names, record, strict=False))
                for record in cursor.fetchall()
                if any(self._has_value(value) for value in record)
            ]
        except sqlite3.Error as exc:
            raise self._parse_error(
                message="Uploaded SQLite database could not be parsed.",
                filename=filename,
                detected_format=ReviewUploadFormat.SQLITE,
            ) from exc
        finally:
            connection.close()

        if not rows:
            raise self._parse_error(
                message="Uploaded SQLite database did not contain any review rows.",
                filename=filename,
                detected_format=ReviewUploadFormat.SQLITE,
            )
        return rows, table_name

    def _rows_from_tabular_values(self, rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
        header_row: tuple[Any, ...] | None = None
        data_rows: list[dict[str, Any]] = []

        for row in rows:
            if not any(self._has_value(cell) for cell in row):
                continue
            if header_row is None:
                header_row = row
                continue

            normalized_row: dict[str, Any] = {}
            for index, header in enumerate(header_row):
                if not self._has_value(header):
                    continue
                value = row[index] if index < len(row) else None
                normalized_row[str(header).strip()] = value
            if any(self._has_value(value) for value in normalized_row.values()):
                data_rows.append(normalized_row)

        return data_rows

    def _build_header_mapping(self, rows: list[dict[str, Any]]) -> dict[str, str]:
        available_headers: dict[str, str] = {}
        for row in rows:
            for key in row:
                normalized = self._normalize_header(key)
                if normalized and normalized not in available_headers:
                    available_headers[normalized] = key

        header_mapping: dict[str, str] = {}
        for canonical in self.CANONICAL_COLUMNS:
            candidate_names = [canonical, *self.COLUMN_ALIASES.get(canonical, [])]
            for candidate_name in candidate_names:
                matched_header = available_headers.get(self._normalize_header(candidate_name))
                if matched_header is not None:
                    header_mapping[canonical] = matched_header
                    break

        missing_columns = [
            column for column in self.REQUIRED_COLUMNS if column not in header_mapping
        ]
        if missing_columns:
            raise AppError(
                code="UPLOAD_COLUMN_MAPPING_ERROR",
                message="Uploaded data is missing required review columns.",
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                details={
                    "missing_columns": missing_columns,
                    "available_columns": sorted(available_headers.values()),
                },
            )
        return header_mapping

    def _build_canonical_rows(
        self,
        rows: list[dict[str, Any]],
        header_mapping: dict[str, str],
    ) -> list[dict[str, Any]]:
        mapped_source_headers = set(header_mapping.values())
        canonical_rows: list[dict[str, Any]] = []

        for row in rows:
            canonical_row: dict[str, Any] = {column: None for column in self.CANONICAL_COLUMNS}
            metadata = self._coerce_metadata(
                row.get(header_mapping["source_metadata"])
            ) if "source_metadata" in header_mapping else {}

            for canonical, source_header in header_mapping.items():
                if canonical == "source_metadata":
                    continue
                canonical_row[canonical] = row.get(source_header)

            for header, value in row.items():
                if header in mapped_source_headers or not self._has_value(value):
                    continue
                metadata.setdefault(header, value)

            canonical_row["source_metadata"] = metadata or None
            if any(
                self._has_value(canonical_row[column])
                for column in self.CANONICAL_COLUMNS
                if column != "source_metadata"
            ):
                canonical_rows.append(canonical_row)

        if not canonical_rows:
            raise AppError(
                code="UPLOAD_PARSE_ERROR",
                message="Uploaded data did not contain any usable review rows after mapping.",
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            )
        return canonical_rows

    def _stage_rows_as_csv(self, rows: list[dict[str, Any]]) -> str:
        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=list(self.CANONICAL_COLUMNS))
        writer.writeheader()
        for row in rows:
            serialized_row = {
                "source_review_id": self._stringify(row.get("source_review_id")),
                "reviewer_name": self._stringify(row.get("reviewer_name")),
                "rating": self._stringify_rating(row.get("rating")),
                "language": self._stringify(row.get("language")),
                "review_text": self._stringify(row.get("review_text")),
                "review_created_at": self._stringify_datetime(row.get("review_created_at")),
                "status": self._stringify(row.get("status")),
                "response_status": self._stringify(row.get("response_status")),
                "source_metadata": json.dumps(
                    row.get("source_metadata") or {},
                    ensure_ascii=True,
                    default=self._json_default,
                )
                if row.get("source_metadata")
                else "",
            }
            writer.writerow(serialized_row)
        return buffer.getvalue()

    def _parse_canonical_csv(self, staged_csv: str) -> list[ReviewImportItem]:
        reader = csv.DictReader(StringIO(staged_csv))
        items: list[ReviewImportItem] = []
        for row_number, row in enumerate(reader, start=2):
            try:
                source_metadata = row.get("source_metadata") or None
                item = ReviewImportItem.model_validate(
                    {
                        "source_review_id": row.get("source_review_id", ""),
                        "reviewer_name": self._blank_to_none(row.get("reviewer_name")),
                        "rating": self._blank_to_none(row.get("rating")),
                        "language": self._blank_to_none(row.get("language")),
                        "review_text": row.get("review_text", ""),
                        "review_created_at": self._blank_to_none(
                            row.get("review_created_at")
                        ),
                        "status": self._blank_to_none(row.get("status")) or "pending",
                        "response_status": self._blank_to_none(
                            row.get("response_status")
                        ),
                        "source_metadata": json.loads(source_metadata)
                        if source_metadata
                        else None,
                    }
                )
            except (ValidationError, json.JSONDecodeError) as exc:
                details = {"row_number": row_number}
                if isinstance(exc, ValidationError):
                    details["errors"] = exc.errors()
                raise AppError(
                    code="UPLOAD_PARSE_ERROR",
                    message="Uploaded data contained an invalid review row.",
                    status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
                    details=details,
                ) from exc
            items.append(item)

        if not items:
            raise AppError(
                code="UPLOAD_PARSE_ERROR",
                message="Uploaded data did not contain any review rows after CSV staging.",
                status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            )
        return items

    def _coerce_metadata(self, value: Any) -> dict[str, Any]:
        if value is None:
            return {}
        if isinstance(value, dict):
            return dict(value)
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                return {}
            try:
                parsed = json.loads(stripped)
            except json.JSONDecodeError:
                return {"raw_metadata": stripped}
            if isinstance(parsed, dict):
                return parsed
            return {"raw_metadata": parsed}
        return {"raw_metadata": value}

    def _normalize_header(self, value: Any) -> str:
        normalized = re.sub(r"[^a-z0-9]+", "_", str(value).strip().lower())
        return normalized.strip("_")

    def _decode_text(self, content: bytes) -> str:
        for encoding in ("utf-8-sig", "utf-8"):
            try:
                return content.decode(encoding)
            except UnicodeDecodeError:
                continue
        raise AppError(
            code="UPLOAD_PARSE_ERROR",
            message="Uploaded text file could not be decoded as UTF-8.",
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
        )

    def _stringify(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, bool):
            return "true" if value else "false"
        return str(value)

    def _stringify_rating(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return ""
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return str(value)
        return self._stringify(value)

    def _stringify_datetime(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return self._stringify(value)

    def _blank_to_none(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, str) and not value.strip():
            return None
        return value

    def _json_default(self, value: Any) -> str:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return str(value)

    def _has_value(self, value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, str):
            return bool(value.strip())
        return True

    def _parse_error(
        self,
        *,
        message: str,
        filename: str,
        detected_format: ReviewUploadFormat,
        details: dict[str, Any] | None = None,
    ) -> AppError:
        error_details = {
            "filename": filename,
            "detected_format": detected_format.value,
        }
        if details:
            error_details.update(details)
        return AppError(
            code="UPLOAD_PARSE_ERROR",
            message=message,
            status_code=status.HTTP_422_UNPROCESSABLE_CONTENT,
            details=error_details,
        )
