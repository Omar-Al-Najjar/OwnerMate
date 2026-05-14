from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
import sqlite3
import unittest
from uuid import uuid4

from openpyxl import Workbook

from backend.app.core.exceptions import AppError
from backend.app.schemas.review import (
    ReviewImportDuplicate,
    ReviewImportResult,
    ReviewRead,
    ReviewUploadImportRequest,
)
from backend.app.services.review_upload import ReviewUploadImportService


class FakeReviewIngestionService:
    def __init__(self) -> None:
        self.last_payload = None
        self.last_input_reference_extra = None
        now = datetime.now(timezone.utc)
        self.sample_review = ReviewRead(
            id=uuid4(),
            business_id=uuid4(),
            review_source_id=None,
            source_type="uploaded",
            source_review_id="ext-1",
            reviewer_name="Jane",
            rating=5,
            language="en",
            review_text="Great service",
            source_metadata={"imported_from": "upload"},
            review_created_at=now,
            ingested_at=now,
            status="pending",
            response_status=None,
            created_at=now,
            updated_at=now,
        )

    def import_reviews(self, payload, *, input_reference_extra=None):
        self.last_payload = payload
        self.last_input_reference_extra = input_reference_extra
        duplicates = []
        if len(payload.reviews) > 1 and payload.reviews[0].source_review_id == payload.reviews[1].source_review_id:
            duplicates = [
                ReviewImportDuplicate(
                    source_review_id=payload.reviews[0].source_review_id,
                    reason="duplicate_in_payload",
                )
            ]

        return (
            ReviewImportResult(
                source=payload.source,
                business_id=payload.business_id,
                review_source_id=payload.review_source_id,
                requested_count=len(payload.reviews),
                imported_count=max(len(payload.reviews) - len(duplicates), 0),
                duplicate_count=len(duplicates),
                processed_count=len(payload.reviews),
                imported_reviews=[self.sample_review],
                duplicates=duplicates,
            ),
            "agent-run-id",
        )


class ReviewUploadImportServiceTests(unittest.TestCase):
    def setUp(self) -> None:
        self.ingestion_service = FakeReviewIngestionService()
        self.service = ReviewUploadImportService(
            review_ingestion_service=self.ingestion_service,
        )
        self.business_id = uuid4()
        self.review_source_id = uuid4()
        self.upload_request = ReviewUploadImportRequest(
            business_id=self.business_id,
            review_source_id=self.review_source_id,
            source="uploaded",
        )

    def test_csv_with_alias_headers_imports_successfully(self) -> None:
        content = (
            "review_id,author,stars,language_code,comment,created_at,metadata,channel\n"
            'ext-1,Jane Doe,5,en,Great service,2026-03-19T09:00:00Z,"{""tag"": ""vip""}",email\n'
        ).encode("utf-8")

        result = self.service.import_reviews(
            upload=self.upload_request,
            filename="reviews.csv",
            content_type="text/csv",
            content=content,
        )

        self.assertEqual(result.requested_count, 1)
        self.assertIsNotNone(self.ingestion_service.last_payload)
        review = self.ingestion_service.last_payload.reviews[0]
        self.assertEqual(review.source_review_id, "ext-1")
        self.assertEqual(review.reviewer_name, "Jane Doe")
        self.assertEqual(review.rating, 5)
        self.assertEqual(review.review_text, "Great service")
        self.assertEqual(
            review.source_metadata,
            {"tag": "vip", "channel": "email"},
        )
        self.assertEqual(
            self.ingestion_service.last_input_reference_extra["upload"]["detected_format"],
            "csv",
        )

    def test_txt_delimiter_detection_supports_common_delimiters(self) -> None:
        samples = {
            ",": "review_id,comment\next-1,Great service\n",
            "\t": "review_id\tcomment\next-1\tGreat service\n",
            "|": "review_id|comment\next-1|Great service\n",
        }

        for delimiter, content in samples.items():
            with self.subTest(delimiter=delimiter):
                self.service.import_reviews(
                    upload=self.upload_request,
                    filename="reviews.txt",
                    content_type="text/plain",
                    content=content.encode("utf-8"),
                )
                self.assertEqual(
                    self.ingestion_service.last_payload.reviews[0].source_review_id,
                    "ext-1",
                )

    def test_xlsx_with_one_non_empty_sheet_imports_successfully(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Reviews"
        sheet.append(["review_id", "comment", "author", "score"])
        sheet.append(["ext-1", "Great service", "Jane", 5])
        content = self._workbook_to_bytes(workbook)

        self.service.import_reviews(
            upload=self.upload_request,
            filename="reviews.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            content=content,
        )

        self.assertEqual(
            self.ingestion_service.last_input_reference_extra["upload"]["selected_source_name"],
            "Reviews",
        )
        self.assertEqual(self.ingestion_service.last_payload.reviews[0].rating, 5)

    def test_xlsx_with_multiple_non_empty_sheets_returns_source_selection_error(self) -> None:
        workbook = Workbook()
        sheet_one = workbook.active
        sheet_one.title = "Reviews"
        sheet_one.append(["review_id", "comment"])
        sheet_one.append(["ext-1", "Great service"])
        sheet_two = workbook.create_sheet("Other")
        sheet_two.append(["review_id", "comment"])
        sheet_two.append(["ext-2", "Slow service"])

        with self.assertRaises(AppError) as context:
            self.service.import_reviews(
                upload=self.upload_request,
                filename="reviews.xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                content=self._workbook_to_bytes(workbook),
            )

        self.assertEqual(context.exception.code, "UPLOAD_SOURCE_SELECTION_ERROR")
        self.assertEqual(context.exception.details["candidate_sources"], ["Reviews", "Other"])

    def test_json_array_and_reviews_object_import_successfully(self) -> None:
        samples = [
            (
                "reviews.json",
                '[{"review_id":"ext-1","comment":"Great service","author":"Jane"}]',
            ),
            (
                "reviews.json",
                '{"reviews":[{"review_id":"ext-2","comment":"Fast service","author":"Ali"}]}',
            ),
        ]

        for filename, content in samples:
            with self.subTest(filename=filename, content=content):
                self.service.import_reviews(
                    upload=self.upload_request,
                    filename=filename,
                    content_type="application/json",
                    content=content.encode("utf-8"),
                )
                self.assertTrue(
                    self.ingestion_service.last_payload.reviews[0].source_review_id.startswith(
                        "ext-"
                    )
                )

    def test_sqlite_with_one_candidate_table_imports_successfully(self) -> None:
        content = self._sqlite_bytes(
            {
                "reviews": {
                    "columns": "review_id TEXT, comment TEXT, author TEXT, score INTEGER",
                    "rows": [("ext-1", "Great service", "Jane", 5)],
                }
            }
        )

        self.service.import_reviews(
            upload=self.upload_request,
            filename="reviews.db",
            content_type="application/octet-stream",
            content=content,
        )

        self.assertEqual(
            self.ingestion_service.last_input_reference_extra["upload"]["selected_source_name"],
            "reviews",
        )
        self.assertEqual(self.ingestion_service.last_payload.reviews[0].review_text, "Great service")

    def test_sqlite_with_multiple_candidate_tables_returns_source_selection_error(self) -> None:
        content = self._sqlite_bytes(
            {
                "reviews_a": {
                    "columns": "review_id TEXT, comment TEXT",
                    "rows": [("ext-1", "Great service")],
                },
                "reviews_b": {
                    "columns": "review_id TEXT, comment TEXT",
                    "rows": [("ext-2", "Slow service")],
                },
            }
        )

        with self.assertRaises(AppError) as context:
            self.service.import_reviews(
                upload=self.upload_request,
                filename="reviews.sqlite",
                content_type="application/octet-stream",
                content=content,
            )

        self.assertEqual(context.exception.code, "UPLOAD_SOURCE_SELECTION_ERROR")
        self.assertEqual(
            context.exception.details["candidate_sources"],
            ["reviews_a", "reviews_b"],
        )

    def test_missing_required_columns_returns_mapping_error(self) -> None:
        content = "author,stars\nJane Doe,5\n".encode("utf-8")

        with self.assertRaises(AppError) as context:
            self.service.import_reviews(
                upload=self.upload_request,
                filename="reviews.csv",
                content_type="text/csv",
                content=content,
            )

        self.assertEqual(context.exception.code, "UPLOAD_COLUMN_MAPPING_ERROR")
        self.assertEqual(
            context.exception.details["missing_columns"],
            ["source_review_id", "review_text"],
        )

    def test_duplicate_rows_flow_through_existing_duplicate_handling(self) -> None:
        content = (
            "review_id,comment\n"
            "dup-1,Great service\n"
            "dup-1,Duplicate row\n"
        ).encode("utf-8")

        result = self.service.import_reviews(
            upload=self.upload_request,
            filename="reviews.csv",
            content_type="text/csv",
            content=content,
        )

        self.assertEqual(len(self.ingestion_service.last_payload.reviews), 2)
        self.assertEqual(result.duplicate_count, 1)
        self.assertEqual(result.duplicates[0].source_review_id, "dup-1")

    def test_invalid_row_values_raise_upload_parse_error(self) -> None:
        content = (
            "review_id,comment,score\n"
            "ext-1,Great service,7\n"
        ).encode("utf-8")

        with self.assertRaises(AppError) as context:
            self.service.import_reviews(
                upload=self.upload_request,
                filename="reviews.csv",
                content_type="text/csv",
                content=content,
            )

        self.assertEqual(context.exception.code, "UPLOAD_PARSE_ERROR")
        self.assertEqual(context.exception.details["row_number"], 2)

    def _workbook_to_bytes(self, workbook: Workbook) -> bytes:
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _sqlite_bytes(self, tables: dict[str, dict[str, object]]) -> bytes:
        connection = sqlite3.connect(":memory:")
        try:
            cursor = connection.cursor()
            for table_name, table_definition in tables.items():
                cursor.execute(
                    f'CREATE TABLE "{table_name}" ({table_definition["columns"]})'
                )
                placeholders = ", ".join(["?"] * len(table_definition["rows"][0]))
                cursor.executemany(
                    f'INSERT INTO "{table_name}" VALUES ({placeholders})',
                    table_definition["rows"],
                )
            connection.commit()
            return connection.serialize()
        finally:
            connection.close()


if __name__ == "__main__":
    unittest.main()
